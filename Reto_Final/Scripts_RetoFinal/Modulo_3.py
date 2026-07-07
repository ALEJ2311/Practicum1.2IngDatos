import pandas as pd
import sqlite3
from openpyxl import load_workbook

def cargar_a_silver(df, nombre_tabla):
    """
    Carga un DataFrame limpio a la base de datos SQLite (Capa Silver).
    Se conecta al mismo archivo creado por los módulos anteriores.
    """
    conexion = sqlite3.connect('macroentorno_silver.db')

    try:
        df.to_sql(nombre_tabla, conexion, if_exists='replace', index=False)
        print(f"Éxito: Tabla '{nombre_tabla}' cargada en macroentorno_silver.db con {len(df)} registros.")
    except Exception as e:
        print(f"Error al cargar la tabla {nombre_tabla}: {e}")
    finally:
        conexion.close()

# Rutas estáticas de los archivos (Bloque 3 — Supercias y MINEDUC)
archivo_companias = 'data/directorio_companias.xlsx'
archivo_ciiu = 'data/bi_ciiu.csv'
archivo_ranking = 'data/bi_ranking.csv'
archivo_amie = 'data/Registro-Administrativo-Historico_2009-202X-Inicio.xlsx'

# ===========================================================================
# SUPERCIAS
# ===========================================================================

# --- 1. Limpieza del Directorio de Compañías ---
print("Procesando Directorio de Compañías...")
# Nota: La lectura de este archivo toma tiempo debido al volumen de datos
df_companias = pd.read_excel(archivo_companias, skiprows=4, engine='openpyxl')

# Estandarización de columnas
df_companias.columns = (df_companias.columns.str.lower()
                        .str.replace(' ', '_')
                        .str.replace('ó', 'o')
                        .str.replace('í', 'i')
                        .str.replace('ñ', 'n'))

# Eliminar registros sin RUC
df_companias = df_companias.dropna(subset=['ruc'])

# Transformación de variable a booleana
if 'presento_balance_inicial' in df_companias.columns:
    df_companias['presento_balance_inicial'] = df_companias['presento_balance_inicial'].map({
        'SI': True,
        'NO': False,
        'NO APLICA': None
    })

cargar_a_silver(df_companias, 'fact_directorio_companias')

# --- 2. Limpieza Clasificador CIIU ---
print("Procesando CIIU...")
df_ciiu = pd.read_csv(archivo_ciiu)

# Estandarización de columnas
df_ciiu.columns = df_ciiu.columns.str.lower().str.replace(' ', '_')

cargar_a_silver(df_ciiu, 'dim_ciiu')

# --- 3. Limpieza del Ranking Empresarial ---
print("Procesando Ranking Empresarial...")
df_ranking = pd.read_csv(archivo_ranking, low_memory=False)

# Renombrar columna de año y estandarizar el resto
if 'AÑO' in df_ranking.columns:
    df_ranking = df_ranking.rename(columns={'AÑO': 'anio'})

df_ranking.columns = df_ranking.columns.str.lower().str.replace(' ', '_')

# Eliminar registros sin RUC
if 'ruc' in df_ranking.columns:
    df_ranking = df_ranking.dropna(subset=['ruc'])

cargar_a_silver(df_ranking, 'fact_ranking_empresarial')

# ===========================================================================
# MINEDUC — AMIE (Registro Administrativo Histórico)
# ===========================================================================

# Hoja con el histórico completo (header en la primera fila, todos los años lectivos).
HOJA_AMIE = 'Historico_Inicio'

# Solo las columnas necesarias para el indicador de bachilleres por provincia.
# El documento pide: filtrar por Nivel de Educación = Bachillerato y el último grado.
# El último grado de bachillerato es "3er Año Bachillerato".
COLUMNAS_AMIE = [
    'Año_lectivo', 'Zona', 'Provincia', 'Cod_Provincia',
    'Canton', 'Cod_Canton', 'Sostenimiento',
    'Total_Estudiantes', 'Bachillerato', '3er Año Bachillerato'
]

# --- 4. Lectura del AMIE ---
print("Procesando MINEDUC / AMIE...")
# El archivo es grande (~300 MB). pd.read_excel carga toda la hoja en memoria y
# resulta muy lento; se lee en modo streaming (read_only) tomando solo las
# columnas necesarias por su posición en la cabecera.
wb = load_workbook(archivo_amie, read_only=True)
ws = wb[HOJA_AMIE]

filas = ws.iter_rows(values_only=True)
cabecera = list(next(filas))
indices = [cabecera.index(c) for c in COLUMNAS_AMIE]  # posiciones de las columnas requeridas

datos = ([fila[i] for i in indices] for fila in filas)
df_amie = pd.DataFrame(datos, columns=COLUMNAS_AMIE)
wb.close()

# --- 5. Estandarización de columnas ---
df_amie = df_amie.rename(columns={
    'Año_lectivo': 'anio_lectivo',
    'Zona': 'zona',
    'Provincia': 'provincia',
    'Cod_Provincia': 'cod_provincia',
    'Canton': 'canton',
    'Cod_Canton': 'cod_canton',
    'Sostenimiento': 'sostenimiento',
    'Total_Estudiantes': 'total_estudiantes',
    'Bachillerato': 'estudiantes_bachillerato',
    '3er Año Bachillerato': 'bachilleres_3ro'
})

# --- 6. Limpieza de nulos ---
# Se eliminan filas sin provincia (posibles notas metodológicas al pie).
df_amie = df_amie.dropna(subset=['provincia'])

# Se descartan zonas no geográficas que distorsionan los cruces provinciales.
df_amie = df_amie[~df_amie['provincia'].isin(['ZONA EN ESTUDIO', 'Zona No Delimitada'])]

# --- 7. Tipos de dato ---
# El año lectivo llega como 'AAAA-AAAA Inicio'; se extrae el año inicial numérico.
df_amie['anio'] = df_amie['anio_lectivo'].astype(str).str.slice(0, 4).astype(int)
for c in ['cod_provincia', 'cod_canton', 'total_estudiantes',
          'estudiantes_bachillerato', 'bachilleres_3ro']:
    df_amie[c] = pd.to_numeric(df_amie[c], errors='coerce').fillna(0).astype(int)

# --- 8. Detalle institucional (grano fino) para trazabilidad ---
cargar_a_silver(df_amie, 'fact_mineduc_amie')

# --- 9. Agregado por provincia y año (insumo directo de la pregunta P3) ---
df_bachilleres = (
    df_amie
    .groupby(['anio', 'provincia', 'cod_provincia'], as_index=False)
    .agg(
        bachilleres_3ro=('bachilleres_3ro', 'sum'),
        estudiantes_bachillerato=('estudiantes_bachillerato', 'sum'),
        total_estudiantes=('total_estudiantes', 'sum'),
        n_instituciones=('provincia', 'size')
    )
    .sort_values(['anio', 'bachilleres_3ro'], ascending=[True, False])
)

cargar_a_silver(df_bachilleres, 'fact_mineduc_bachilleres')

print("\nModulo 3 finalizado. Bloque 3 (Supercias + MINEDUC) cargado exitosamente.")